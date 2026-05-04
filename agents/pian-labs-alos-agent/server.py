from __future__ import annotations

import base64
import csv
import io
import json
import math
import os
from pathlib import Path
from typing import Any
from urllib.parse import quote
from urllib.request import Request as UrlRequest, urlopen

from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse
from pydantic import BaseModel
try:
    from pydantic import ConfigDict
except Exception:  # pragma: no cover
    ConfigDict = None  # type: ignore

import sys

ROOT = Path(__file__).resolve().parent
EC2_ROOT = ROOT.parents[1]
sys.path.insert(0, str(EC2_ROOT))

from ec2_shared.agent_response import as_text, card, failed, needs_input, require_fields, success
from ec2_shared.ui import render_agent_window

AGENT_ID = "pian-labs-alos-agent"
AGENT_NAME = "Pian Labs ALOS Agent"

DEFAULT_ENTITIES: dict[str, list[dict[str, Any]]] = {
    "shipments": [
        {"id": "SHP-1001", "origin": "Delhi", "destination": "Mumbai", "status": "in_transit", "priority": "high"},
        {"id": "SHP-1002", "origin": "Bengaluru", "destination": "Hyderabad", "status": "planned", "priority": "normal"},
    ],
    "vehicles": [
        {"id": "TRK-22", "type": "truck", "status": "available", "location": "Delhi"},
        {"id": "VAN-09", "type": "van", "status": "maintenance", "location": "Bengaluru"},
    ],
    "warehouses": [
        {"id": "WH-DEL", "city": "Delhi", "capacity": 12000, "used": 9200},
        {"id": "WH-MUM", "city": "Mumbai", "capacity": 10000, "used": 8100},
    ],
}


class ActionRequest(BaseModel):
    action: str | None = None
    if ConfigDict:
        model_config = ConfigDict(extra="allow")
    else:
        class Config:
            extra = "allow"


def _payload(model: ActionRequest) -> dict[str, Any]:
    if hasattr(model, "model_dump"):
        return model.model_dump()
    return model.dict()


def _data_url_to_bytes(value: str) -> bytes:
    if "," in value and value.startswith("data:"):
        value = value.split(",", 1)[1]
    return base64.b64decode(value)


def _json_path(data: Any, path: str | None) -> Any:
    if not path:
        return data
    current = data
    for part in path.strip("$.").split("."):
        if not part:
            continue
        if isinstance(current, dict):
            current = current.get(part)
        elif isinstance(current, list) and part.isdigit():
            current = current[int(part)]
        else:
            return []
    return current


def _entity_name(payload: dict[str, Any]) -> str:
    return as_text(payload.get("entity") or payload.get("collection") or payload.get("table") or "shipments").replace(" ", "_").lower()


def _registry_from_rows(rows_by_entity: dict[str, list[dict[str, Any]]]) -> list[dict[str, Any]]:
    registry = []
    for entity, rows in rows_by_entity.items():
        fields = sorted({key for row in rows for key in row.keys()})
        registry.append({"entity": entity, "fields": fields, "sample_count": len(rows), "writable": True})
    return registry


def _matches(row: dict[str, Any], filters: dict[str, Any]) -> bool:
    for key, expected in filters.items():
        if expected is None or expected == "":
            continue
        actual = row.get(key)
        if isinstance(expected, str):
            if str(actual).lower() != expected.lower():
                return False
        elif actual != expected:
            return False
    return True


def _haversine_km(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    radius = 6371.0
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    delta_phi = math.radians(lat2 - lat1)
    delta_lambda = math.radians(lon2 - lon1)
    a = math.sin(delta_phi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(delta_lambda / 2) ** 2
    return 2 * radius * math.atan2(math.sqrt(a), math.sqrt(1 - a))


def _fetch_json(url: str, timeout: int = 12) -> Any:
    request = UrlRequest(url, headers={"User-Agent": "pian-alos-agent/1.0"})
    with urlopen(request, timeout=timeout) as response:
        return json.loads(response.read().decode("utf-8"))


def _geocode(place: str) -> dict[str, Any] | None:
    data = _fetch_json(f"https://geocoding-api.open-meteo.com/v1/search?name={quote(place)}&count=1&language=en&format=json")
    results = data.get("results") or []
    if not results:
        return None
    item = results[0]
    return {"name": item.get("name"), "country": item.get("country"), "lat": float(item["latitude"]), "lon": float(item["longitude"])}


def _weather_for(lat: float, lon: float) -> dict[str, Any]:
    data = _fetch_json(f"https://api.open-meteo.com/v1/forecast?latitude={lat}&longitude={lon}&current=temperature_2m,precipitation,wind_speed_10m,weather_code")
    return data.get("current") or {}


def _route_distance(origin: dict[str, Any], destination: dict[str, Any]) -> dict[str, Any]:
    coords = f"{origin['lon']},{origin['lat']};{destination['lon']},{destination['lat']}"
    try:
        data = _fetch_json(f"https://router.project-osrm.org/route/v1/driving/{coords}?overview=false")
        route = (data.get("routes") or [])[0]
        return {"distance_km": round(route.get("distance", 0) / 1000, 1), "duration_min": round(route.get("duration", 0) / 60), "source": "osrm"}
    except Exception:
        return {"distance_km": round(_haversine_km(origin["lat"], origin["lon"], destination["lat"], destination["lon"]), 1), "duration_min": None, "source": "great_circle"}


def _verdict(weather: dict[str, Any]) -> str:
    wind = float(weather.get("wind_speed_10m") or 0)
    precipitation = float(weather.get("precipitation") or 0)
    if wind >= 45 or precipitation >= 8:
        return "reroute"
    if wind >= 28 or precipitation >= 2:
        return "caution"
    return "go"


async def _mongo_collection(entity: str):
    try:
        from motor.motor_asyncio import AsyncIOMotorClient
    except Exception:
        return None
    uri = os.getenv("MONGODB_URI", "mongodb://localhost:27017")
    db_name = os.getenv("MONGODB_DB", "alos")
    client = AsyncIOMotorClient(uri, serverSelectionTimeoutMS=1500)
    try:
        await client.admin.command("ping")
    except Exception:
        client.close()
        return None
    return client[db_name][entity]


async def _load_rows(entity: str, limit: int = 100) -> list[dict[str, Any]]:
    collection = await _mongo_collection(entity)
    if collection is None:
        return list(DEFAULT_ENTITIES.get(entity, []))[:limit]
    rows = await collection.find({}, {"_id": 0}).limit(limit).to_list(length=limit)
    return rows


async def _describe(payload: dict[str, Any], action: str) -> dict[str, Any]:
    registry_env = as_text(os.getenv("ALOS_ENTITY_REGISTRY_JSON"))
    if registry_env:
        try:
            registry = json.loads(registry_env)
        except Exception:
            registry = []
    else:
        rows_by_entity = {entity: await _load_rows(entity, 20) for entity in DEFAULT_ENTITIES}
        registry = _registry_from_rows(rows_by_entity)
    return success(
        agent=AGENT_ID,
        action=action,
        summary=f"ALOS workspace has {len(registry)} registered entities.",
        result={"registry": registry},
        cards=[card("Entity registry", "Workspace entities and fields are available.", {item["entity"]: ", ".join(item.get("fields", [])) for item in registry})],
    )


async def _summarize(payload: dict[str, Any], action: str) -> dict[str, Any]:
    entity = _entity_name(payload)
    rows = await _load_rows(entity, int(payload.get("limit") or 500))
    if not rows:
        return needs_input(agent=AGENT_ID, action=action, message=f"No rows are available for entity '{entity}'. Onboard data or choose another entity.", missing_fields=["entity"])
    status_counts: dict[str, int] = {}
    for row in rows:
        status = as_text(row.get("status") or row.get("state") or "unknown")
        status_counts[status] = status_counts.get(status, 0) + 1
    fields = sorted({key for row in rows for key in row.keys()})
    return success(
        agent=AGENT_ID,
        action=action,
        summary=f"{entity} has {len(rows)} sampled records and {len(fields)} fields.",
        result={"entity": entity, "count": len(rows), "fields": fields, "status_counts": status_counts, "sample": rows[:5]},
        cards=[card(f"{entity} summary", "Entity summary created.", {"records": len(rows), "fields": len(fields), **status_counts})],
    )


async def _list_records(payload: dict[str, Any], action: str) -> dict[str, Any]:
    entity = _entity_name(payload)
    limit = int(payload.get("limit") or payload.get("count") or 25)
    filters = payload.get("filters") if isinstance(payload.get("filters"), dict) else {}
    rows = [row for row in await _load_rows(entity, 500) if _matches(row, filters)][: max(1, min(limit, 100))]
    return success(
        agent=AGENT_ID,
        action=action,
        summary=f"Loaded {len(rows)} {entity} records.",
        result={"entity": entity, "records": rows, "filters": filters},
        cards=[card(f"{entity} records", f"{len(rows)} rows returned.", {"filters": json.dumps(filters) if filters else "none"}), *[card(as_text(row.get("id") or row.get("name") or "Record"), "", row) for row in rows[:5]]],
    )


async def _create_record(payload: dict[str, Any], action: str) -> dict[str, Any]:
    entity = _entity_name(payload)
    record = payload.get("record") if isinstance(payload.get("record"), dict) else None
    if not record:
        return needs_input(agent=AGENT_ID, action=action, message="Provide a record object to insert.", missing_fields=["record"])
    collection = await _mongo_collection(entity)
    if collection is None:
        return needs_input(agent=AGENT_ID, action=action, message="MongoDB is required to create persistent records.", missing_fields=["MONGODB_URI"])
    result = await collection.insert_one(record)
    return success(agent=AGENT_ID, action=action, summary=f"Created one {entity} record.", result={"entity": entity, "inserted_id": str(result.inserted_id), "record": record}, cards=[card("Record created", "MongoDB insert completed.", {"entity": entity, "inserted_id": str(result.inserted_id)})])


async def _update_records(payload: dict[str, Any], action: str) -> dict[str, Any]:
    entity = _entity_name(payload)
    filters = payload.get("filters") if isinstance(payload.get("filters"), dict) else None
    updates = payload.get("updates") if isinstance(payload.get("updates"), dict) else None
    missing = []
    if not filters:
        missing.append("filters")
    if not updates:
        missing.append("updates")
    if missing:
        return needs_input(agent=AGENT_ID, action=action, message="ALOS needs filters and updates to modify records safely.", missing_fields=missing)
    collection = await _mongo_collection(entity)
    if collection is None:
        return needs_input(agent=AGENT_ID, action=action, message="MongoDB is required to update persistent records.", missing_fields=["MONGODB_URI"])
    result = await collection.update_many(filters, {"$set": updates})
    return success(agent=AGENT_ID, action=action, summary=f"Updated {result.modified_count} {entity} records.", result={"entity": entity, "matched": result.matched_count, "modified": result.modified_count}, cards=[card("Records updated", "MongoDB update completed.", {"matched": result.matched_count, "modified": result.modified_count})])


async def _route_weather(payload: dict[str, Any], action: str) -> dict[str, Any]:
    missing = require_fields(AGENT_ID, action, payload, ["origin", "destination"])
    if missing:
        return missing
    origin_name = as_text(payload.get("origin"))
    destination_name = as_text(payload.get("destination"))
    origin = _geocode(origin_name)
    destination = _geocode(destination_name)
    if not origin or not destination:
        return needs_input(agent=AGENT_ID, action=action, message="ALOS could not geocode one of the route endpoints.", missing_fields=["origin", "destination"])
    distance = _route_distance(origin, destination)
    weather_points = [
        {"point": "origin", "place": origin, "weather": _weather_for(origin["lat"], origin["lon"])},
        {"point": "destination", "place": destination, "weather": _weather_for(destination["lat"], destination["lon"])},
    ]
    verdicts = [_verdict(item["weather"]) for item in weather_points]
    verdict = "reroute" if "reroute" in verdicts else "caution" if "caution" in verdicts else "go"
    return success(
        agent=AGENT_ID,
        action=action,
        summary=f"Route verdict from {origin_name} to {destination_name}: {verdict}.",
        result={"origin": origin, "destination": destination, "route": distance, "weather_points": weather_points, "verdict": verdict},
        cards=[
            card("Route verdict", f"{origin_name} to {destination_name}: {verdict}", {"distance_km": distance["distance_km"], "duration_min": distance["duration_min"], "source": distance["source"]}),
            *[card(f"Weather at {item['place']['name']}", "", item["weather"]) for item in weather_points],
        ],
        next_actions=["Proceed with dispatch" if verdict == "go" else "Review alternate route or delay dispatch"],
    )


async def _onboard_csv(payload: dict[str, Any], action: str) -> dict[str, Any]:
    entity = _entity_name(payload)
    data = as_text(payload.get("csv") or payload.get("file_data_url") or payload.get("data"))
    if not data:
        return needs_input(agent=AGENT_ID, action=action, message="Upload a CSV file or paste CSV text.", missing_fields=["csv"])
    if data.startswith("data:"):
        data = _data_url_to_bytes(data).decode("utf-8-sig")
    rows = list(csv.DictReader(io.StringIO(data)))
    collection = await _mongo_collection(entity)
    inserted = 0
    if collection is not None and rows:
        result = await collection.insert_many(rows)
        inserted = len(result.inserted_ids)
    return success(agent=AGENT_ID, action=action, summary=f"Onboarded {len(rows)} CSV rows for {entity}.", result={"entity": entity, "rows": rows[:20], "row_count": len(rows), "inserted": inserted}, cards=[card("CSV onboarded", "Rows were parsed and inserted when MongoDB was available.", {"entity": entity, "rows": len(rows), "inserted": inserted})])


async def _onboard_excel(payload: dict[str, Any], action: str) -> dict[str, Any]:
    data = as_text(payload.get("excel_base64") or payload.get("file_data_url"))
    if not data:
        return needs_input(agent=AGENT_ID, action=action, message="Upload an Excel workbook.", missing_fields=["file_data_url"])
    try:
        from openpyxl import load_workbook
    except Exception:
        return needs_input(agent=AGENT_ID, action=action, message="The Excel importer needs openpyxl installed in this service.", missing_fields=["openpyxl"])

    workbook = load_workbook(io.BytesIO(_data_url_to_bytes(data)), data_only=True)
    imported: dict[str, int] = {}
    samples: dict[str, list[dict[str, Any]]] = {}
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [as_text(cell).replace(" ", "_").lower() for cell in rows[0]]
        headers = [header or f"field_{index + 1}" for index, header in enumerate(headers)]
        records = [
            {headers[index]: value for index, value in enumerate(row) if index < len(headers)}
            for row in rows[1:]
            if any(value not in (None, "") for value in row)
        ]
        entity = sheet.title.replace(" ", "_").lower()
        samples[entity] = records[:10]
        imported[entity] = len(records)
        collection = await _mongo_collection(entity)
        if collection is not None and records:
            await collection.insert_many(records)

    if not imported:
        return needs_input(agent=AGENT_ID, action=action, message="The Excel workbook did not contain importable rows.", missing_fields=["workbook_rows"])
    return success(
        agent=AGENT_ID,
        action=action,
        summary=f"Onboarded {sum(imported.values())} Excel rows across {len(imported)} entities.",
        result={"imported": imported, "samples": samples},
        cards=[card("Excel onboarded", "Each sheet was converted into a logistics entity.", imported)],
    )


async def _pull_rest_api(payload: dict[str, Any], action: str) -> dict[str, Any]:
    missing = require_fields(AGENT_ID, action, payload, ["url"])
    if missing:
        return missing
    url = as_text(payload.get("url"))
    headers = payload.get("headers") if isinstance(payload.get("headers"), dict) else {}
    request = UrlRequest(url, headers={str(k): str(v) for k, v in headers.items()})
    with urlopen(request, timeout=20) as response:
        data = json.loads(response.read().decode("utf-8"))
    records = _json_path(data, as_text(payload.get("json_path") or payload.get("path")) or None)
    if isinstance(records, dict):
        records = [records]
    if not isinstance(records, list):
        records = []
    return success(agent=AGENT_ID, action=action, summary=f"Pulled {len(records)} records from REST API.", result={"url": url, "records": records[:50], "count": len(records)}, cards=[card("REST API import", f"{len(records)} records returned.", {"url": url})])


async def _clone_mongo(payload: dict[str, Any], action: str) -> dict[str, Any]:
    missing = require_fields(AGENT_ID, action, payload, ["source_mongodb_uri", "source_database"])
    if missing:
        return missing
    try:
        from motor.motor_asyncio import AsyncIOMotorClient
    except Exception:
        return needs_input(agent=AGENT_ID, action=action, message="The Mongo clone importer needs motor installed in this service.", missing_fields=["motor"])

    source_uri = as_text(payload.get("source_mongodb_uri"))
    source_db_name = as_text(payload.get("source_database"))
    requested_collections = payload.get("collections")
    if isinstance(requested_collections, str):
        collection_names = [item.strip() for item in requested_collections.split(",") if item.strip()]
    elif isinstance(requested_collections, list):
        collection_names = [as_text(item) for item in requested_collections if as_text(item)]
    else:
        collection_names = []

    source_client = AsyncIOMotorClient(source_uri, serverSelectionTimeoutMS=2500)
    local_uri = os.getenv("MONGODB_URI", "mongodb://localhost:27017")
    local_db_name = os.getenv("MONGODB_DB", "alos")
    local_client = AsyncIOMotorClient(local_uri, serverSelectionTimeoutMS=2500)
    try:
        await source_client.admin.command("ping")
        await local_client.admin.command("ping")
        source_db = source_client[source_db_name]
        if not collection_names:
            collection_names = await source_db.list_collection_names()
        imported: dict[str, int] = {}
        for collection_name in collection_names:
            rows = await source_db[collection_name].find({}).limit(int(payload.get("limit") or 1000)).to_list(length=int(payload.get("limit") or 1000))
            clean_rows = []
            for row in rows:
                row.pop("_id", None)
                clean_rows.append(row)
            if clean_rows:
                await local_client[local_db_name][collection_name].insert_many(clean_rows)
            imported[collection_name] = len(clean_rows)
        return success(
            agent=AGENT_ID,
            action=action,
            summary=f"Cloned {sum(imported.values())} records from remote Mongo into ALOS.",
            result={"source_database": source_db_name, "imported": imported},
            cards=[card("Mongo cloned", "Remote collections were copied into the local ALOS database.", imported)],
        )
    finally:
        source_client.close()
        local_client.close()


async def _start_empty(payload: dict[str, Any], action: str) -> dict[str, Any]:
    collection_status: dict[str, int] = {}
    for entity, rows in DEFAULT_ENTITIES.items():
        collection = await _mongo_collection(entity)
        if collection is not None:
            await collection.insert_many(rows)
        collection_status[entity] = len(rows)
    registry = _registry_from_rows(DEFAULT_ENTITIES)
    return success(
        agent=AGENT_ID,
        action=action,
        summary="Started ALOS with the built-in logistics schema.",
        result={"registry": registry, "seeded": collection_status},
        cards=[card("Empty workspace started", "Built-in logistics entities are ready for smoke tests.", collection_status)],
        next_actions=["Describe workspace", "Summarize shipments", "Check route weather"],
    )


async def _run_alos(payload: dict[str, Any], action: str) -> dict[str, Any]:
    prompt = as_text(payload.get("prompt") or payload.get("message") or payload.get("query") or payload.get("parameters"))
    if not prompt:
        return needs_input(agent=AGENT_ID, action=action, message="Tell ALOS what logistics operation to perform.", missing_fields=["prompt"])
    lower = prompt.lower()
    if "weather" in lower or "route" in lower or ("from" in lower and " to " in lower):
        if not payload.get("origin") or not payload.get("destination"):
            parts = prompt.split(" to ", 1)
            if len(parts) == 2:
                origin = parts[0].split(" from ")[-1].strip(" .")
                destination = parts[1].strip(" .")
                payload = {**payload, "origin": payload.get("origin") or origin, "destination": payload.get("destination") or destination}
        return await _route_weather(payload, "check_route_weather")
    if "summar" in lower or "overview" in lower:
        return await _summarize(payload, "summarize_entity")
    if "list" in lower or "show" in lower:
        return await _list_records(payload, "list_records")
    return success(
        agent=AGENT_ID,
        action=action,
        summary="ALOS can proceed, but it needs a specific logistics tool selection for this request.",
        result={"prompt": prompt, "suggested_actions": ["describe_workspace", "summarize_entity", "list_records", "check_route_weather"]},
        cards=[card("Clarification", "Choose the logistics action that best matches the task.", {"suggested_actions": "describe_workspace, summarize_entity, list_records, check_route_weather"})],
        next_actions=["Describe workspace", "Summarize an entity", "Check route weather"],
    )


CAPABILITIES = [
    {"name": "run_alos", "label": "Run ALOS", "description": "Route a natural-language logistics request to the right ALOS tool.", "required": ["prompt"], "optional": ["entity", "origin", "destination"]},
    {"name": "describe_workspace", "label": "Describe Workspace", "description": "List logistics entities and their fields.", "required": [], "optional": []},
    {"name": "summarize_entity", "label": "Summarize Entity", "description": "Summarize totals, fields, statuses, and sample records.", "required": ["entity"], "optional": ["limit"]},
    {"name": "list_records", "label": "List Records", "description": "List records with optional filters.", "required": ["entity"], "optional": ["filters", "limit"]},
    {"name": "create_record", "label": "Create Record", "description": "Insert one writable MongoDB record.", "required": ["entity", "record"], "optional": []},
    {"name": "update_records", "label": "Update Records", "description": "Update MongoDB records using explicit filters.", "required": ["entity", "filters", "updates"], "optional": []},
    {"name": "check_route_weather", "label": "Check Route Weather", "description": "Geocode a route, fetch OSRM distance, sample Open-Meteo weather, and return go/caution/reroute.", "required": ["origin", "destination"], "optional": []},
    {"name": "onboard_csv", "label": "Onboard CSV", "description": "Import CSV text or uploaded CSV into a logistics entity.", "required": ["entity", "csv"], "optional": ["file_data_url"]},
    {"name": "onboard_excel", "label": "Onboard Excel", "description": "Import an Excel workbook with one entity per sheet.", "required": ["file_data_url"], "optional": []},
    {"name": "pull_rest_api", "label": "Pull REST API", "description": "Fetch records from a REST API and optional JSON path.", "required": ["url"], "optional": ["headers", "json_path"]},
    {"name": "clone_mongo", "label": "Clone Mongo", "description": "Clone selected collections from a remote MongoDB into the local ALOS workspace.", "required": ["source_mongodb_uri", "source_database"], "optional": ["collections", "limit"]},
    {"name": "start_empty", "label": "Start Empty", "description": "Seed the built-in logistics schema for an empty workspace.", "required": [], "optional": []},
    {"name": "ask_user", "label": "Ask User", "description": "Pause and request missing logistics data.", "required": ["question"], "optional": []},
    {"name": "list_capabilities", "label": "List Capabilities", "description": "Show ALOS capabilities.", "required": [], "optional": []},
]

UI_SPEC = {
    "name": AGENT_NAME,
    "description": "Autonomous Logistics Operating System for logistics workspace onboarding, entity summaries, record operations, and weather-aware routing.",
    "endpoint": "/alos/action",
    "actions": CAPABILITIES,
    "examples": [
        "Describe the current logistics workspace.",
        "Summarize shipments and highlight delayed items.",
        "Check route weather from Delhi to Mumbai before dispatch.",
    ],
    "scope": [
        "MongoDB-first logistics workspace, CSV/REST onboarding, entity registry, records, and route weather.",
        "Open-Meteo and OSRM are used only for route checks.",
    ],
    "usage": [
        "Use onboarding actions before asking about your own logistics data.",
        "Use ask_user-style missing-field cards when route or entity data is ambiguous.",
        "Use check_route_weather before dispatch decisions.",
    ],
}

app = FastAPI(title=AGENT_NAME, version="1.0.0")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_credentials=True, allow_methods=["*"], allow_headers=["*"])


@app.get("/", response_class=HTMLResponse)
async def root() -> str:
    return render_agent_window(UI_SPEC)


@app.get("/alos/health")
async def health() -> dict[str, Any]:
    return {"status": "ok", "agent": AGENT_ID, "actions": [item["name"] for item in CAPABILITIES]}


@app.post("/alos/action")
async def action(request: ActionRequest) -> dict[str, Any]:
    payload = _payload(request)
    selected = as_text(payload.get("action") or "run_alos")
    try:
        if selected == "run_alos":
            return await _run_alos(payload, selected)
        if selected == "describe_workspace":
            return await _describe(payload, selected)
        if selected == "summarize_entity":
            return await _summarize(payload, selected)
        if selected == "list_records":
            return await _list_records(payload, selected)
        if selected == "create_record":
            return await _create_record(payload, selected)
        if selected == "update_records":
            return await _update_records(payload, selected)
        if selected == "check_route_weather":
            return await _route_weather(payload, selected)
        if selected == "onboard_csv":
            return await _onboard_csv(payload, selected)
        if selected == "onboard_excel":
            return await _onboard_excel(payload, selected)
        if selected == "pull_rest_api":
            return await _pull_rest_api(payload, selected)
        if selected == "clone_mongo":
            return await _clone_mongo(payload, selected)
        if selected == "start_empty":
            return await _start_empty(payload, selected)
        if selected == "ask_user":
            question = as_text(payload.get("question")) or "What logistics detail should ALOS use?"
            return needs_input(agent=AGENT_ID, action=selected, message=question, missing_fields=[as_text(payload.get("field") or "answer")])
        if selected == "list_capabilities":
            return success(agent=AGENT_ID, action=selected, summary="ALOS capabilities loaded.", result={"actions": CAPABILITIES}, cards=[card("Capabilities", "ALOS supports workspace description, onboarding, record operations, and route weather checks.", {"actions": ", ".join(item["name"] for item in CAPABILITIES)})])
        return needs_input(agent=AGENT_ID, action=selected, message=f"ALOS does not expose the action '{selected}'.", missing_fields=["action"])
    except Exception as exc:
        return failed(agent=AGENT_ID, action=selected, public_message="ALOS could not complete this logistics request yet.", error=exc)


if __name__ == "__main__":
    import uvicorn

    uvicorn.run("server:app", host="0.0.0.0", port=8054)
